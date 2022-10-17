import csv
from openpyxl import load_workbook, Workbook

from importer import *

class FilesImporter:

    def __init__(self, filePath, returnColumns = None, defaultHeaders=None, defaultDataTypes=None, rowDataType=ROW_TYPE_DICT, noneStrings=DEFAULT_NONE_STRINGS, noneExceptions=None, rowLimit=None):
        self.defaultHeaders = defaultHeaders
        self.defaultDataTypes = defaultDataTypes
        self.rowDataType = rowDataType
        self.noneStrings = noneStrings
        self.noneExceptions = noneExceptions

        # transform into for loop that will extract reader, headers and data for each of the three dataframes

        self.fileReader = self.getFileReader(filePath)
        self.headers = self.processHeaders(list(next(self.fileReader))) # Get first row from file reader as header
        self.data = self.getData(rowLimit)

        # Select specific columns
        self.data = self.selectColumns(self.data, returnColumns)
        self.headers = returnColumns
        print(f'Headers: {self.headers}')

        print('done')

    def printRows(self, numRows): # Method to help troubleshooting by simply printing rows
        for idx, row in enumerate(self.data):
            print(row)
            if idx == numRows - 1:
                break


    def getFileReader(self, filePath):
        """ Get an iterator used to get each row in a CSV or XLSX file
        """

        if f'.{FILE_TYPE_CSV}' in filePath:
            self.file = open(filePath, encoding="utf8")
            self.fileType = FILE_TYPE_CSV
            return csv.reader(self.file)
        elif f'.{FILE_TYPE_XLS}' in filePath:
            self.file = load_workbook(filePath, data_only=True)
            worksheet = self.file.active
            self.fileType = FILE_TYPE_XLS
            return worksheet.iter_rows(values_only=True)
        else:
            raise(ValueError('Unsupported file type'))


    def processHeaders(self, headers):
        # If there's no default return given headers
        if not self.defaultHeaders:
            return headers

        # If there is replace by index/value
        newHeaders = []
        for idx, header in enumerate(headers):
            headerByIdx = self.defaultHeaders.get(idx)
            headerByVal = self.defaultHeaders.get(header)
            if headerByIdx:
                newHeaders.append(headerByIdx)
            elif headerByVal:
                newHeaders.append(headerByVal)
            else:
                newHeaders.append(header)
        return newHeaders


    def getData(self, rowLimit):
        # get all data from file import and return list of processed records
        data = []
        for idx, row in enumerate(self.fileReader):
            if rowLimit and idx == rowLimit: # Break out if reach row limit
                break
            processedRow = self.processRow(row) # Call processRow on each row

            # Print type of each column (for each header). Zip only if row datatype is not already dict
            if idx == 0:
                headersAndVals = zip(self.headers, processedRow) if self.rowDataType != ROW_TYPE_DICT else processedRow.items()
                for header, val in headersAndVals:
                    print(f'{header} type is {type(val)}')
            data.append(processedRow)

            # close CSV manually
        if self.fileType == FILE_TYPE_CSV:
            self.file.close()

        return data

    def processRow(self, row):
        # process row into specified data type uses processValue
        processedRow = []
        # For each header, value get the default data type for that header and translate the value
        for header, val in zip(self.headers, row):

            defaultDataType = self.defaultDataTypes.get(header) if self.defaultDataTypes else None
            processedRow.append(self.processValue(val, header, defaultDataType))
        if self.rowDataType == ROW_TYPE_LIST:
            return processedRow
        if self.rowDataType == ROW_TYPE_DICT:
            return {header: val for header, val in zip(self.headers, processedRow)}
        if self.rowDataType == ROW_TYPE_TUPLE:
            return tuple(processedRow)

    def processValue(self, val, header, defaultDataType):
        # Processes a value. hangles None, null strings and converting via custom function
        if val is None:
            return val
        if isinstance(val, str):
            val = val.strip()
            if val.lower() in self.noneStrings:
                if self.noneExceptions and header == self.noneExceptions:
                    return defaultDataType(val)
                return None
        if defaultDataType:
            val = defaultDataType(val)
        return val

    # \\ Refactor this to be within the processing. ideally input is just the list of headers and if header is not specified it's removed
    def selectColumns(self, data, returnColumns):
        if returnColumns is None:
            return data

        selectData = []
        if self.rowDataType == ROW_TYPE_DICT:
            for row in data:
                # Save new dict row if header is in returnColumns
                newRow = {header: val for header, val in row.items() if header in returnColumns}
                selectData.append(newRow)
        elif self.rowDataType == ROW_TYPE_LIST:
            # Generate index of each inputted column header
            keepIdx = []
            for col in returnColumns:
                keepIdx.append(self.headers.index(col))
            for row in data:
                newRow = list(row[idx] for idx in keepIdx)

                selectData.append(newRow)

        # [elif tuple]

        return selectData
